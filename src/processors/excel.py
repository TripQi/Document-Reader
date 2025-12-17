# Excel 文档处理器
"""处理 Excel (XLS/XLSX) 和 ODS 文档"""

import os
from typing import Any, Optional

from .base import DocumentProcessor
from ..models import (
    ReadDocumentInput,
    DocumentResult,
    DocumentMetadata,
    ContentItem,
)
from ..utils import (
    validate_file,
    format_file_size,
    handle_file_error,
    determine_excel_output_format,
    PREVIEW_ROWS,
)


class ExcelProcessor(DocumentProcessor):
    """Excel 文档处理器

    支持 XLS、XLSX、ODS 格式。
    智能格式转换策略:
    - 行数 ≤ 50: 转换为 Markdown 表格（易读）
    - 行数 > 50: 转换为 CSV 格式（紧凑）
    - 预览模式: 仅读取前 10 行
    """

    @property
    def supported_extensions(self) -> list[str]:
        """返回支持的文件扩展名列表"""
        return ['.xls', '.xlsx', '.ods']

    def supports_extension(self, ext: str) -> bool:
        """检查是否支持该文件扩展名"""
        return ext.lower() in self.supported_extensions

    async def process(self, params: ReadDocumentInput) -> DocumentResult:
        """处理 Excel 文档

        Args:
            params: 读取文档的输入参数

        Returns:
            DocumentResult: 包含文件名、类型、元数据和内容的结构化结果
        """
        file_path = params.file_path

        # 验证文件
        validate_file(file_path)

        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == '.xlsx':
                return await self._process_xlsx(params)
            elif ext == '.xls':
                return await self._process_xls(params)
            elif ext == '.ods':
                return await self._process_ods(params)
            else:
                raise ValueError(f"不支持的文件格式: {ext}")

        except Exception as e:
            error_msg = handle_file_error(e, file_path)
            raise ValueError(error_msg) from e

    async def _process_xlsx(self, params: ReadDocumentInput) -> DocumentResult:
        """处理 XLSX 文件"""
        import openpyxl

        file_path = params.file_path

        # 加载工作簿
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        try:
            # 获取工作表
            sheet_names = wb.sheetnames
            sheet = self._get_sheet(wb, params.sheet_name, params.sheet_index, sheet_names)
            sheet_name = sheet.title

            # 读取数据
            data, total_rows, total_cols = self._read_xlsx_data(
                sheet,
                params.preview_mode,
                params.max_rows
            )

            # 确定输出格式
            format_hint = determine_excel_output_format(
                total_rows,
                params.preview_mode,
                params.max_rows
            )

            # 构建元数据
            metadata = DocumentMetadata(
                file_name=os.path.basename(file_path),
                file_type="Excel",
                file_size=format_file_size(os.path.getsize(file_path)),
                sheet_name=sheet_name,
                total_rows=total_rows,
                total_columns=total_cols,
                available_sheets=sheet_names,
                format_hint=format_hint,
                preview_mode=params.preview_mode,
            )

            # 构建内容
            content = []
            if data:
                content.append(ContentItem(
                    type="table",
                    data=data
                ))

            return DocumentResult(
                file_name=os.path.basename(file_path),
                file_type="Excel",
                metadata=metadata,
                content=content,
                format_hint=format_hint
            )

        finally:
            wb.close()

    def _get_sheet(
        self,
        wb: Any,
        sheet_name: Optional[str],
        sheet_index: Optional[int],
        sheet_names: list[str]
    ) -> Any:
        """获取指定的工作表

        Args:
            wb: 工作簿对象
            sheet_name: 工作表名称
            sheet_index: 工作表索引
            sheet_names: 可用工作表名称列表

        Returns:
            工作表对象

        Raises:
            ValueError: 工作表不存在
        """
        if sheet_name:
            if sheet_name not in sheet_names:
                available = ', '.join(sheet_names)
                raise ValueError(
                    f"错误: 工作表 '{sheet_name}' 不存在。可用工作表: {available}"
                )
            return wb[sheet_name]

        if sheet_index is not None:
            if sheet_index < 0 or sheet_index >= len(sheet_names):
                raise ValueError(
                    f"错误: 工作表索引 {sheet_index} 超出范围。"
                    f"有效范围: 0-{len(sheet_names) - 1}"
                )
            return wb[sheet_names[sheet_index]]

        # 默认返回第一个工作表
        return wb[sheet_names[0]]

    def _read_xlsx_data(
        self,
        sheet: Any,
        preview_mode: bool,
        max_rows: Optional[int]
    ) -> tuple[list[list[Any]], int, int]:
        """读取 XLSX 工作表数据

        Args:
            sheet: 工作表对象
            preview_mode: 是否为预览模式
            max_rows: 最大行数限制

        Returns:
            tuple: (数据列表, 总行数, 总列数)
        """
        # 获取实际数据范围
        total_rows = sheet.max_row or 0
        total_cols = sheet.max_column or 0

        # 确定读取行数
        if preview_mode:
            rows_to_read = min(PREVIEW_ROWS, total_rows)
        elif max_rows:
            rows_to_read = min(max_rows, total_rows)
        else:
            rows_to_read = total_rows

        # 读取数据
        data: list[list[Any]] = []
        for row_idx, row in enumerate(sheet.iter_rows(max_row=rows_to_read), 1):
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    row_data.append("")
                else:
                    row_data.append(str(value))
            data.append(row_data)

        return data, total_rows, total_cols

    async def _process_xls(self, params: ReadDocumentInput) -> DocumentResult:
        """处理 XLS 文件"""
        import xlrd

        file_path = params.file_path

        # 打开工作簿
        wb = xlrd.open_workbook(file_path)

        try:
            # 获取工作表
            sheet_names = wb.sheet_names()
            sheet = self._get_xls_sheet(wb, params.sheet_name, params.sheet_index, sheet_names)
            sheet_name = sheet.name

            # 读取数据
            data, total_rows, total_cols = self._read_xls_data(
                sheet,
                params.preview_mode,
                params.max_rows
            )

            # 确定输出格式
            format_hint = determine_excel_output_format(
                total_rows,
                params.preview_mode,
                params.max_rows
            )

            # 构建元数据
            metadata = DocumentMetadata(
                file_name=os.path.basename(file_path),
                file_type="Excel",
                file_size=format_file_size(os.path.getsize(file_path)),
                sheet_name=sheet_name,
                total_rows=total_rows,
                total_columns=total_cols,
                available_sheets=sheet_names,
                format_hint=format_hint,
                preview_mode=params.preview_mode,
            )

            # 构建内容
            content = []
            if data:
                content.append(ContentItem(
                    type="table",
                    data=data
                ))

            return DocumentResult(
                file_name=os.path.basename(file_path),
                file_type="Excel",
                metadata=metadata,
                content=content,
                format_hint=format_hint
            )

        finally:
            # xlrd 不需要显式关闭
            pass

    def _get_xls_sheet(
        self,
        wb: Any,
        sheet_name: Optional[str],
        sheet_index: Optional[int],
        sheet_names: list[str]
    ) -> Any:
        """获取 XLS 工作表"""
        if sheet_name:
            if sheet_name not in sheet_names:
                available = ', '.join(sheet_names)
                raise ValueError(
                    f"错误: 工作表 '{sheet_name}' 不存在。可用工作表: {available}"
                )
            return wb.sheet_by_name(sheet_name)

        if sheet_index is not None:
            if sheet_index < 0 or sheet_index >= len(sheet_names):
                raise ValueError(
                    f"错误: 工作表索引 {sheet_index} 超出范围。"
                    f"有效范围: 0-{len(sheet_names) - 1}"
                )
            return wb.sheet_by_index(sheet_index)

        return wb.sheet_by_index(0)

    def _read_xls_data(
        self,
        sheet: Any,
        preview_mode: bool,
        max_rows: Optional[int]
    ) -> tuple[list[list[Any]], int, int]:
        """读取 XLS 工作表数据"""
        total_rows = sheet.nrows
        total_cols = sheet.ncols

        # 确定读取行数
        if preview_mode:
            rows_to_read = min(PREVIEW_ROWS, total_rows)
        elif max_rows:
            rows_to_read = min(max_rows, total_rows)
        else:
            rows_to_read = total_rows

        # 读取数据
        data: list[list[Any]] = []
        for row_idx in range(rows_to_read):
            row_data = []
            for col_idx in range(total_cols):
                value = sheet.cell_value(row_idx, col_idx)
                if value == '':
                    row_data.append("")
                else:
                    row_data.append(str(value))
            data.append(row_data)

        return data, total_rows, total_cols

    async def _process_ods(self, params: ReadDocumentInput) -> DocumentResult:
        """处理 ODS 文件"""
        from odf import text as odf_text
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell

        file_path = params.file_path

        # 加载文档
        doc = load(file_path)

        # 获取所有表格
        tables = doc.getElementsByType(Table)
        sheet_names = [t.getAttribute('name') for t in tables]

        if not tables:
            raise ValueError(f"错误: ODS 文件 '{file_path}' 中没有找到工作表。")

        # 选择工作表
        table = self._get_ods_table(tables, params.sheet_name, params.sheet_index, sheet_names)
        sheet_name = table.getAttribute('name')

        # 读取数据
        data, total_rows, total_cols = self._read_ods_data(
            table,
            params.preview_mode,
            params.max_rows
        )

        # 确定输出格式
        format_hint = determine_excel_output_format(
            total_rows,
            params.preview_mode,
            params.max_rows
        )

        # 构建元数据
        metadata = DocumentMetadata(
            file_name=os.path.basename(file_path),
            file_type="ODS",
            file_size=format_file_size(os.path.getsize(file_path)),
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_columns=total_cols,
            available_sheets=sheet_names,
            format_hint=format_hint,
            preview_mode=params.preview_mode,
        )

        # 构建内容
        content = []
        if data:
            content.append(ContentItem(
                type="table",
                data=data
            ))

        return DocumentResult(
            file_name=os.path.basename(file_path),
            file_type="ODS",
            metadata=metadata,
            content=content,
            format_hint=format_hint
        )

    def _get_ods_table(
        self,
        tables: list,
        sheet_name: Optional[str],
        sheet_index: Optional[int],
        sheet_names: list[str]
    ) -> Any:
        """获取 ODS 表格"""
        if sheet_name:
            for table in tables:
                if table.getAttribute('name') == sheet_name:
                    return table
            available = ', '.join(sheet_names)
            raise ValueError(
                f"错误: 工作表 '{sheet_name}' 不存在。可用工作表: {available}"
            )

        if sheet_index is not None:
            if sheet_index < 0 or sheet_index >= len(tables):
                raise ValueError(
                    f"错误: 工作表索引 {sheet_index} 超出范围。"
                    f"有效范围: 0-{len(tables) - 1}"
                )
            return tables[sheet_index]

        return tables[0]

    def _read_ods_data(
        self,
        table: Any,
        preview_mode: bool,
        max_rows: Optional[int]
    ) -> tuple[list[list[Any]], int, int]:
        """读取 ODS 表格数据"""
        from odf.table import TableRow, TableCell
        from odf import text as odf_text

        rows = table.getElementsByType(TableRow)
        total_rows = len(rows)

        # 确定读取行数
        if preview_mode:
            rows_to_read = min(PREVIEW_ROWS, total_rows)
        elif max_rows:
            rows_to_read = min(max_rows, total_rows)
        else:
            rows_to_read = total_rows

        # 读取数据
        data: list[list[Any]] = []
        max_cols = 0

        for row_idx, row in enumerate(rows[:rows_to_read]):
            row_data = []
            cells = row.getElementsByType(TableCell)

            for cell in cells:
                # 获取单元格文本
                text_content = ""
                for p in cell.getElementsByType(odf_text.P):
                    text_content += "".join(
                        node.data for node in p.childNodes
                        if hasattr(node, 'data')
                    )
                row_data.append(text_content)

            if row_data:
                data.append(row_data)
                max_cols = max(max_cols, len(row_data))

        return data, total_rows, max_cols
